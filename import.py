import os
import openpyxl
import sqlite3
import pandas as pd


def import_products():
    """Import products data data from xlsx using pandas"""
    connect = sqlite3.connect('food_db.sqlite3')
    wb = pd.read_excel('data/products.xlsx', sheet_name=None)

    for sheet in wb:
        wb[sheet].to_sql('calculator_products', if_exists='append', con=connect, index=False)
    connect.commit()
    connect.close()


def import_categories() -> None:
    """Import food categories data from xlsx using openpyxl"""
    project_dir = os.path.abspath(os.path.curdir)
    base_name = 'food_db.sqlite3'

    # Соединение с базой данных
    connect = sqlite3.connect(project_dir + '/' + base_name)
    cursor = connect.cursor()

    # Работа с таблицей XLSX
    file = openpyxl.load_workbook('data/categories.xlsx', data_only=True)
    sheet = file['Sheet1']
    # Объявление списка

    for row in range(2, sheet.max_row + 1):
        data = []
        # Цикл по столбцам от 1 до 4
        for col in range(1, 4):
            # value содержит значение ячейки с координатами row col
            value = sheet.cell(row, col).value
            # Список который мы потом будем добавлять
            data.append(value)
        cursor.execute("INSERT INTO calculator_categories VALUES (?, ?, ?);", (data[0], data[1], data[2]))

    # Сохраняем изменения
    connect.commit()
    # Закрытие соединения
    connect.close()


if __name__ == '__main__':
    iscategories = False
    isproducts = False

    if iscategories:
        import_categories()

    if isproducts:
        import_products()
